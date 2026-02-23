import pyreadstat
import pandas as pd
import math
from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side)
from openpyxl.utils import get_column_letter

df, meta = pyreadstat.read_sav(r'D:\SPSS_PS\Employee data.sav')
df['gender'] = df['gender'].map({'m': 'Male', 'f': 'Female'})

wb = Workbook()

# ── Styles ──────────────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")
SUBHEAD_FILL  = PatternFill("solid", fgColor="2E75B6")
ALT_FILL      = PatternFill("solid", fgColor="DEEAF1")
TOTAL_FILL    = PatternFill("solid", fgColor="BDD7EE")
WHITE_FILL    = PatternFill("solid", fgColor="FFFFFF")

H_FONT  = Font(bold=True, color="FFFFFF", size=11)
SH_FONT = Font(bold=True, color="FFFFFF", size=10)
TOT_FONT= Font(bold=True, size=10)
BODY    = Font(size=10)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center")

def thin_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def style_row(ws, row, col_start, col_end, fill, font, align=CENTER, border=True):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill  = fill
        cell.font  = font
        cell.alignment = align
        if border:
            cell.border = thin_border()

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — Variable Information
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Variable Info"
ws1.sheet_view.showGridLines = False

# Title
ws1.merge_cells("A1:F1")
ws1["A1"] = "SPSS Employee Dataset — Variable Information"
ws1["A1"].fill  = HEADER_FILL
ws1["A1"].font  = Font(bold=True, color="FFFFFF", size=13)
ws1["A1"].alignment = CENTER
ws1.row_dimensions[1].height = 30

# Sub-info
ws1.merge_cells("A2:F2")
ws1["A2"] = "Source: Employee data.sav  |  474 employees  |  10 variables  |  1 missing value (bdate)"
ws1["A2"].fill = SUBHEAD_FILL
ws1["A2"].font = Font(color="FFFFFF", size=10, italic=True)
ws1["A2"].alignment = CENTER
ws1.row_dimensions[2].height = 18

ws1.append([])  # blank row

# Headers
headers = ["#", "Variable", "Label", "Type", "Value Labels", "Missing"]
ws1.append(headers)
for c, h in enumerate(headers, 1):
    cell = ws1.cell(row=4, column=c)
    cell.value = h
    cell.fill  = SUBHEAD_FILL
    cell.font  = SH_FONT
    cell.alignment = CENTER
    cell.border = thin_border()
ws1.row_dimensions[4].height = 22

# Variable rows
var_info = [
    (1,  "id",       "Employee Code",              "Numeric",  "—",                              "None"),
    (2,  "gender",   "Gender",                     "String",   "f = Female, m = Male",            "None"),
    (3,  "bdate",    "Date of Birth",              "Date",     "—",                              "1 missing"),
    (4,  "educ",     "Educational Level (years)",  "Numeric",  "8 to 21 years",                   "None"),
    (5,  "jobcat",   "Employment Category",        "Numeric",  "1=Clerical, 2=Custodial, 3=Manager", "None"),
    (6,  "salary",   "Current Salary",             "Numeric",  "—",                              "None"),
    (7,  "salbegin", "Beginning Salary",           "Numeric",  "—",                              "None"),
    (8,  "jobtime",  "Months since Hire",          "Numeric",  "—",                              "None"),
    (9,  "prevexp",  "Previous Experience (months)","Numeric", "—",                              "None"),
    (10, "minority", "Minority Classification",    "Numeric",  "0 = No, 1 = Yes",                "None"),
]

for i, row_data in enumerate(var_info):
    r = 5 + i
    fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
    for c, val in enumerate(row_data, 1):
        cell = ws1.cell(row=r, column=c, value=val)
        cell.fill = fill
        cell.font = BODY
        cell.alignment = LEFT if c > 1 else CENTER
        cell.border = thin_border()
    ws1.row_dimensions[r].height = 18

set_col_widths(ws1, [5, 12, 30, 10, 38, 10])

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — Summary Statistics
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Summary Statistics")
ws2.sheet_view.showGridLines = False

ws2.merge_cells("A1:G1")
ws2["A1"] = "SPSS Employee Dataset — Summary Statistics"
ws2["A1"].fill = HEADER_FILL
ws2["A1"].font = Font(bold=True, color="FFFFFF", size=13)
ws2["A1"].alignment = CENTER
ws2.row_dimensions[1].height = 30

ws2.merge_cells("A2:G2")
ws2["A2"] = "Numeric variables only  |  N = 474"
ws2["A2"].fill = SUBHEAD_FILL
ws2["A2"].font = Font(color="FFFFFF", size=10, italic=True)
ws2["A2"].alignment = CENTER
ws2.row_dimensions[2].height = 18

ws2.append([])

stat_headers = ["Variable", "N", "Mean", "Std Dev", "Min", "Median", "Max"]
ws2.append(stat_headers)
for c, h in enumerate(stat_headers, 1):
    cell = ws2.cell(row=4, column=c)
    cell.value = h
    cell.fill  = SUBHEAD_FILL
    cell.font  = SH_FONT
    cell.alignment = CENTER
    cell.border = thin_border()
ws2.row_dimensions[4].height = 22

num_cols = ["salary", "salbegin", "educ", "jobtime", "prevexp"]
labels   = ["Current Salary", "Beginning Salary", "Education (yrs)", "Months since Hire", "Prev Experience (mo)"]
money    = {"salary", "salbegin"}

for i, (col, lbl) in enumerate(zip(num_cols, labels)):
    r = 5 + i
    fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
    s = df[col]
    def fmt(v):
        return f"${v:,.0f}" if col in money else f"{v:,.1f}"
    row_vals = [lbl, int(s.count()), fmt(s.mean()), fmt(s.std()), fmt(s.min()), fmt(s.median()), fmt(s.max())]
    for c, val in enumerate(row_vals, 1):
        cell = ws2.cell(row=r, column=c, value=val)
        cell.fill = fill
        cell.font = BODY
        cell.alignment = CENTER
        cell.border = thin_border()
    ws2.row_dimensions[r].height = 18

# Gender breakdown
ws2.append([])
r = 11
ws2.merge_cells(f"A{r}:G{r}")
ws2[f"A{r}"] = "Gender Distribution"
ws2[f"A{r}"].fill = SUBHEAD_FILL
ws2[f"A{r}"].font = SH_FONT
ws2[f"A{r}"].alignment = CENTER
ws2.row_dimensions[r].height = 20

for j, (label, count) in enumerate(df['gender'].value_counts().items()):
    r2 = 12 + j
    fill = ALT_FILL if j % 2 == 0 else WHITE_FILL
    pct = count / len(df) * 100
    for c, val in enumerate([label, count, f"{pct:.1f}%", "", "", "", ""], 1):
        cell = ws2.cell(row=r2, column=c, value=val)
        cell.fill = fill
        cell.font = BODY
        cell.alignment = CENTER
        cell.border = thin_border()
    ws2.row_dimensions[r2].height = 18

set_col_widths(ws2, [22, 8, 14, 14, 14, 14, 14])

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — Beginning Salary by Gender & Education
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Salary by Gender & Education")
ws3.sheet_view.showGridLines = False

ws3.merge_cells("A1:G1")
ws3["A1"] = "Beginning Salary by Gender and Education"
ws3["A1"].fill = HEADER_FILL
ws3["A1"].font = Font(bold=True, color="FFFFFF", size=13)
ws3["A1"].alignment = CENTER
ws3.row_dimensions[1].height = 30

ws3.merge_cells("A2:G2")
ws3["A2"] = "Mean beginning salary (USD) and employee count by education level and gender"
ws3["A2"].fill = SUBHEAD_FILL
ws3["A2"].font = Font(color="FFFFFF", size=10, italic=True)
ws3["A2"].alignment = CENTER
ws3.row_dimensions[2].height = 18

ws3.append([])

# Merged gender column headers
ws3.merge_cells("B4:C4")
ws3["B4"] = "Female"
ws3.merge_cells("D4:E4")
ws3["D4"] = "Male"
ws3.merge_cells("F4:G4")
ws3["F4"] = "Total"
for col in ["B", "D", "F"]:
    ws3[f"{col}4"].fill = SUBHEAD_FILL
    ws3[f"{col}4"].font = SH_FONT
    ws3[f"{col}4"].alignment = CENTER
    ws3[f"{col}4"].border = thin_border()
for col in ["C", "E", "G"]:
    ws3[f"{col}4"].fill = SUBHEAD_FILL
    ws3[f"{col}4"].border = thin_border()
ws3.row_dimensions[4].height = 20

sub_headers = ["Education (yrs)", "Mean Salary", "N", "Mean Salary", "N", "Mean Salary", "N"]
ws3.append(sub_headers)
for c, h in enumerate(sub_headers, 1):
    cell = ws3.cell(row=5, column=c)
    cell.value = h
    cell.fill  = SUBHEAD_FILL
    cell.font  = SH_FONT
    cell.alignment = CENTER
    cell.border = thin_border()
ws3.row_dimensions[5].height = 22

pivot = df.pivot_table(values='salbegin', index='educ', columns='gender',
                        aggfunc=['mean','count'], margins=True, margins_name='Total')

for i, educ_val in enumerate(pivot.index):
    r = 6 + i
    is_total = (educ_val == 'Total')
    fill = TOTAL_FILL if is_total else (ALT_FILL if i % 2 == 0 else WHITE_FILL)
    font = TOT_FONT if is_total else BODY
    label = 'Total' if is_total else str(int(educ_val))

    def g(stat, gender):
        try:
            v = pivot.loc[educ_val, (stat, gender)]
            return v if not math.isnan(v) else None
        except: return None

    mf, nf = g('mean','Female'), g('count','Female')
    mm, nm = g('mean','Male'),   g('count','Male')
    mt, nt = g('mean','Total'),  g('count','Total')

    row_vals = [
        label,
        f"${mf:,.0f}" if mf else "—", int(nf) if nf else "—",
        f"${mm:,.0f}" if mm else "—", int(nm) if nm else "—",
        f"${mt:,.0f}" if mt else "—", int(nt) if nt else "—",
    ]
    for c, val in enumerate(row_vals, 1):
        cell = ws3.cell(row=r, column=c, value=val)
        cell.fill = fill
        cell.font = font
        cell.alignment = CENTER
        cell.border = thin_border()
    ws3.row_dimensions[r].height = 18

# Key findings note
note_row = 6 + len(pivot.index) + 1
ws3.merge_cells(f"A{note_row}:G{note_row}")
ws3[f"A{note_row}"] = "💡 Key Finding: Male employees earned $7,209 more on average than females in beginning salary. The gap widens significantly at 16+ years of education."
ws3[f"A{note_row}"].fill = PatternFill("solid", fgColor="FFF2CC")
ws3[f"A{note_row}"].font = Font(size=10, italic=True, color="7F6000")
ws3[f"A{note_row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws3.row_dimensions[note_row].height = 32

set_col_widths(ws3, [18, 14, 8, 14, 8, 14, 8])

# ── Save ─────────────────────────────────────────────────────────────────────
out = r'D:\SPSS_PS\Employee_Data_Analysis.xlsx'
wb.save(out)
print(f"Saved: {out}")
